import json
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from openpyxl.styles import PatternFill, Border, Side

riolinks = [
    'https://www.raidbots.com/simbot/report/3KfRVLXkSay5NLExPg5wow', #Stronk
    'https://www.raidbots.com/simbot/report/657pGSZ85NPVKi4xhiRmM3', #Lass
]

workbook = Workbook()
sheet = workbook.active

names = []
mean_dps = []
item_ids = {}
item_levels = {}
dps_increase = {}
listed_items = []

for link in riolinks:
    response = requests.get(link+'/data.json')

    if response.status_code == 200:
        data = response.json()
        items = data['simbot']['meta']['rawFormData']['droptimizerItems']
        for item in items:
            item_id = item['id']
            item_name = item['item']['name']
            item_level = item['item']['itemLevel']
            item_tags = item['item'].get('tags', [])
            if 'catalyst' not in item_tags:
                item_ids[item_id] = item_name
                item_levels[item_id] = item_level

        # Find player name
        player_name = data['sim']['players'][0]['name']

        # Calculate DPS for each item
        results = data['sim']['profilesets']['results']
        items_calculated = []
        player_dps_increase = {}
        for result in results:
            item_id = result['name']
            dps_gain = result['mean']
            item_name = item_ids.get(item_id)
            if item_name:
                if item_name not in items_calculated:
                    player_dps_increase[item_name] = dps_gain   
                    items_calculated.append(item_name)
                    dps_increase[player_name] = player_dps_increase
                else:
                    if (dps_gain > dps_increase[player_name].get(item_name)):
                        player_dps_increase[item_name] = dps_gain
                        dps_increase[player_name] = player_dps_increase

        #Pre sim dps
        player_mean_dps = data['sim']['players'][0]['collected_data']['dps']['mean']
        mean_dps.append(player_mean_dps)

#add names and ilvl
sheet.cell(row=1, column=1, value='Item').alignment = Alignment(horizontal='left')
for row, (item_id, item_name) in enumerate(item_ids.items(), start=2):
    item_level = item_levels[item_id]
    item_text = f"{item_name} ({item_level})"
    if (item_name not in listed_items):
        sheet.cell(row=row, column=1, value=item_text).alignment = Alignment(horizontal='left')
        listed_items.append(item_name)

#Fill headers
sheet.insert_cols(2, 3)
sheet.cell(row=1, column=2, value='1st').alignment = Alignment(horizontal='left')
sheet.cell(row=1, column=3, value='2nd').alignment = Alignment(horizontal='left')
sheet.cell(row=1, column=4, value='3rd').alignment = Alignment(horizontal='left')

#Find best gains
for row, item_name in enumerate(item_ids.values(), start=2):
    dps_gains = [] 
    for col, player_name in enumerate(dps_increase.keys(), start=5):
        dps_gain = dps_increase[player_name].get(item_name)
        if (dps_gain is not None):
            rounded_dps_gain = round(dps_gain - mean_dps[col-5], 0)
            sheet.cell(row=row, column=col, value=rounded_dps_gain).alignment = Alignment(horizontal='left')
            dps_gains.append((rounded_dps_gain, player_name))

    dps_gains.sort(reverse=True)

    #Fill best dps
    for i, (dps_gain, player_name) in enumerate(dps_gains[:3], start=2):
        dps_gain_text = f"({dps_gain}) {player_name}"
        cell = sheet.cell(row=row, column=i)
        sheet.cell(row=row, column=i, value=dps_gain_text).alignment = Alignment(horizontal='left')
        if dps_gain is not None:
            if dps_gain < 0:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            elif dps_gain >= 200:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

    for col in range(5, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        value = cell.value
        if value is not None:
            if value < 0:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            elif value >= 200:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))

#Fill player names
for col, player_name in enumerate(dps_increase.keys(), start=5):
    sheet.cell(row=1, column=col, value=player_name).alignment = Alignment(horizontal='left')

#Style column size
for column in sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column[0].column_letter].width = adjusted_width

for row in range(sheet.max_row, 0, -1):
    cell_value = sheet.cell(row=row, column=1).value
    if cell_value is None:
        sheet.delete_rows(row, 1)

workbook.save('teamsims.xlsx')
